import requests
import json
import openpyxl
import channels
import time
from openpyxl import load_workbook
headers = {
  'user-agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36',
  'authority' : 'api.slackarchive.io',
  'Remote-Adress' : '51.15.69.64:443',
  'referer' : 'https://productmanagerhq.slackarchive.io/ama/page-9',
  'x-alt-referer' : 'https://productmanagerhq.slackarchive.io'
}
proxies = {
"http" : 'http://122.13.15.9:80'
}

k=0
i=7969
wb = load_workbook(filename = 'slackyarchive2.xlsx')
ws = wb['Sheet1']
dictlist = [{'id' :'C064NNJF3','name' :'newtopm'}, {'id' :'C084MTKEW','name' :'productfeedback'}, {'id' :'C0B277J5S','name' :'uxresearch'}, {'id' :'C0EFXRDUK','name' :'writing'}, {'id' :'C0EG09UDP','name' :'dailystandup'}, {'id' :'C26S4UVL6','name' :'podcasts'}, {'id' :'C03PSJ0MG','name' :'general'}, {'id' :'C0NHKRXFC','name' :'ama'}, {'id' :'C04UTQ1FX','name' :'pminterviews'}, {'id' :'C04N1D366','name' :'pmjobs'}, {'id' :'C04PGAP4T','name' :'pmresources'}]

for pepe in dictlist:
    k=0
    while(1):
        #print(pepe.get('id'))

        try:
            time.sleep(5)
            url = 'https://api.slackarchive.io/v1/messages?size=100&team=T03PSJ0M8&channel=' + pepe.get('id') + '&offset=' + str(k)
            r = requests.get(url, headers=headers)
            strin = '{"messages": [],'
            #print(strin)
            #print(strin)
            cont = r.content.decode("utf-8")
            #print(len("{'messages': [], 'total': 30, 'aggs': {'buckets': {}}, 'related': {'users': {}}})"))
            #print(cont)
            print(url)
            if(cont.find('Too many requests') == 1):
                print(r.content)
                print("Error")
                time.sleep(10)
                r = requests.get(url, headers=headers)
            #{'messages': [], 'total': 30, 'aggs': {'buckets': {}}, 'related': {'users': {}}}
            elif((len(cont))<101):
                print(r.content)
                print("Error 2")

                raise EnvironmentError

            #print(r.content)
            con = str(r.content)
            kek = (json.loads(r.content))
            #print(kek)
        #A = SITE, B=CHANNEL, C= AUTHOR, D=MESSAGE, E=TIMESTAMP
            length = len(kek.get('messages'))

            for listic in kek.get('messages'):
                #print('saving')

                channel = pepe.get('name')
                ws['A' + str(i)] = 'productmanagerhq'
                ws['B' + str(i)] = channel
                ws['C' + str(i)] = listic.get('user')
                ws['D' + str(i)] = listic.get('text')
                ws['E' + str(i)] = listic.get('ts')
                i+=1

            k+=100

        except:
            #print("kek")
            break
    wb.save(filename = 'slackyarchive2.xlsx')



#https://api.slackarchive.io/v1/messages?size=100&team=T03PSJ0M8&channel=C0NHKRXFC&offset=1400


#print(kek.get('messages')[0].get('text'))
